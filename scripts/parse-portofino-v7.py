#!/usr/bin/env python3
"""
Parsea la hoja PRESUPUESTO del Excel V7 de Portofino y deja un JSON con las
partidas listas para cargar. NO toca la base de datos — solo lee el Excel.

Convención de carga (pedida por MJ):
  - costo total de la partida  = col I (P.U) x col H (cantidad)  -> unitPrice/total
  - mano de obra               = col L (P.U MO)                  -> costLabor (POR UNIDAD)
  - todo el resto (sin desglose fino) -> costMaterial = P.U - P.U MO (POR UNIDAD)

Los 13 "rojos" (filas 95-107) no tienen costo al cliente: van con noCobrado=True
y su valor es la mano de obra (lo que se le paga al maestro).
"""
import json
import sys
import unicodedata
import openpyxl

XLSX = "/Users/mjblanco/Library/CloudStorage/GoogleDrive-mjblanco@blarq.cl/Unidades compartidas/BLARQ-SOCIOS/3- PROYECTOS TERMINADOS/2026/60_PORTOFINO 4208 (ANGELICA OVALLE)/5- PRESUPUESTO/V7_ENTREGA/V7_ OBRA_PORTOFINO_17.06.26.xlsx"
SALIDA = "/private/tmp/claude-501/-Users-mjblanco-Desktop-blarq-app/322f2d89-4cde-4d41-9f3e-919656befb73/scratchpad/portofino-v7-partidas.json"

FILA_INI, FILA_FIN = 21, 107   # 21 = "1 DEMOLICION", 107 = último rojo
COL = dict(item=3, nombre=4, desc=6, unidad=7, cant=8, pu=9, total=10, pu_mo=12, total_mo=13)


def slug(nombre: str) -> str:
    """Misma forma de clave legacy que usó la carga de V6: minúsculas, sin tildes, con _."""
    s = unicodedata.normalize("NFKD", nombre).encode("ascii", "ignore").decode()
    return s.lower().strip().replace(" ", "_")


def num(v):
    return float(v) if isinstance(v, (int, float)) else 0.0


def main():
    wb = openpyxl.load_workbook(XLSX, data_only=True)
    ws = wb["PRESUPUESTO"]

    capitulos = []      # {numero, nombre}
    partidas = []
    cap_actual = None
    sub_actual = None

    for r in range(FILA_INI, FILA_FIN + 1):
        item = ws.cell(r, COL["item"]).value
        nombre = ws.cell(r, COL["nombre"]).value
        if nombre is None:
            continue
        nombre = str(nombre).strip()
        item = str(item).strip() if item is not None else ""

        # Encabezado de capítulo: item es un entero solo ("1", "2", ... "8")
        if item.isdigit():
            cap_actual = {"numero": int(item), "nombre": nombre}
            capitulos.append(cap_actual)
            sub_actual = None
            continue

        cant = num(ws.cell(r, COL["cant"]).value)
        pu = num(ws.cell(r, COL["pu"]).value)
        total = num(ws.cell(r, COL["total"]).value)
        pu_mo = num(ws.cell(r, COL["pu_mo"]).value)
        total_mo = num(ws.cell(r, COL["total_mo"]).value)
        unidad_raw = ws.cell(r, COL["unidad"]).value
        unidad = unidad_raw or "GL"
        desc = ws.cell(r, COL["desc"]).value

        # Sub-capítulo (COCINA / BAÑOS): tiene item con 2 niveles pero ni costo
        # ni cantidad ni unidad. Es un separador visual, no una partida.
        if pu == 0 and total == 0 and total_mo == 0 and cant == 0 and unidad_raw is None:
            sub_actual = nombre
            partidas.append({
                "fila": r, "tipo": "subcapitulo", "capitulo": cap_actual["numero"],
                "item": item, "nombre": nombre,
            })
            continue

        # "Rojo": sin costo al cliente, solo mano de obra -> BLARQ lo absorbe.
        rojo = (total == 0 and pu == 0 and total_mo > 0)

        partidas.append({
            "fila": r,
            "tipo": "rojo" if rojo else "partida",
            "capitulo": cap_actual["numero"],
            "subCapitulo": sub_actual,
            "item": item,
            "nombre": nombre,
            "descripcion": str(desc).strip() if desc else None,
            "unidad": str(unidad).strip(),
            "cantidad": cant,
            # Rojo: el "precio" de la partida ES la mano de obra (lo que se paga
            # al maestro). No suma al cliente porque va con noCobrado.
            "puCosto": pu_mo if rojo else pu,
            "totalCosto": total_mo if rojo else total,
            "puMO": pu_mo,
            "totalMO": total_mo,
            "puMaterial": (0.0 if rojo else pu - pu_mo),
            "noCobrado": rojo,
        })

    # ── Controles ────────────────────────────────────────────────────────────
    reales = [p for p in partidas if p["tipo"] != "subcapitulo"]
    cobrables = [p for p in reales if not p["noCobrado"]]
    rojos = [p for p in reales if p["noCobrado"]]

    cd = sum(p["totalCosto"] for p in cobrables)
    mo = sum(p["totalMO"] for p in reales)
    mo_rojos = sum(p["totalMO"] for p in rojos)

    ctrl = {
        "costoDirecto": cd,
        "costoDirectoExcel": ws["J108"].value,
        "moTotal": mo,
        "moTotalExcel": ws["M115"].value,
        "moRojos": mo_rojos,
        "gg": cd * 0.20,
        "utilidad": cd * 0.07,
        "neto": cd * 1.27,
        "iva": cd * 1.27 * 0.19,
        "totalConIva": cd * 1.27 * 1.19,
        "totalConIvaExcel": ws["J113"].value,
        "nPartidas": len(reales),
        "nCobrables": len(cobrables),
        "nRojos": len(rojos),
        "nSubcapitulos": len(partidas) - len(reales),
    }

    # Descuadres partida por partida (P.U - MO deberia ser >= 0)
    negativos = [p for p in cobrables if p["puMaterial"] < -0.5]

    out = {"capitulos": capitulos, "partidas": partidas, "control": ctrl,
           "puMaterialNegativo": [(p["item"], p["nombre"], p["puMaterial"]) for p in negativos]}
    with open(SALIDA, "w") as f:
        json.dump(out, f, indent=2, ensure_ascii=False)

    print("CAPITULOS:", " | ".join(f'{c["numero"]} {c["nombre"]}' for c in capitulos))
    print()
    for k, v in ctrl.items():
        print(f"  {k:22} {v:,.2f}" if isinstance(v, float) else f"  {k:22} {v:,}")
    print()
    print("dif costo directo vs Excel:", round(cd - ctrl["costoDirectoExcel"], 2))
    print("dif MO vs Excel:", round(mo - ctrl["moTotalExcel"], 2))
    print("dif total c/IVA vs Excel:", round(ctrl["totalConIva"] - ctrl["totalConIvaExcel"], 2))
    print("partidas con P.U material negativo (MO > costo):", negativos)
    print("\nJSON ->", SALIDA)


if __name__ == "__main__":
    sys.exit(main())
